import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/tmp/sheet.xlsx')
if 'MANUAL MONTHLY TEMPLATE' in wb.sheetnames:
    sheet = wb['MANUAL MONTHLY TEMPLATE']
    
    widths = {}
    for i in range(1, 10):
        col_letter = get_column_letter(i)
        dim = sheet.column_dimensions[col_letter]
        widths[i] = dim.width if dim.width else 'default'
        
    cell = sheet.cell(row=1, col=1)
    font = cell.font
    alignment = cell.alignment
    fill = cell.fill
    
    print('Widths:', widths)
    print('Font Name:', font.name, 'Size:', font.size, 'Bold:', font.bold, 'Italic:', font.italic, 'Color:', font.color.rgb if font.color else font.color)
    print('Alignment H:', alignment.horizontal, 'V:', alignment.vertical, 'WrapText:', alignment.wrapText)
    print('BgColor:', getattr(fill.start_color, 'rgb', None) if fill and fill.start_color else None)
else:
    print("Sheet not found.")
